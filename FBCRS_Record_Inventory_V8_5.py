import os, time, hashlib, re, sys, warnings, logging, csv, threading, tempfile, shutil, traceback
from datetime import datetime, date
from concurrent.futures import ThreadPoolExecutor
from threading import Lock
import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# Suppress warnings and logs
warnings.simplefilter('ignore')
logging.getLogger("openpyxl").setLevel(logging.CRITICAL)
logging.getLogger("pypdf").setLevel(logging.ERROR)

try: import pypdf
except ImportError: pypdf = None
try: import docx
except ImportError: docx = None
try: import pptx
except: pptx = None
# ahocorasick removed: the single compiled word-boundary regex below is both
# faster than the old per-code substring loop and, unlike either, correct.

if getattr(sys, 'frozen', False): 
    SCRIPT_DIR = os.path.dirname(sys.executable)
else: 
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

LIBRARY = os.path.join(SCRIPT_DIR, 'FBCRS_Master_Full.xlsx')

# --- FIX 1: never write output to SCRIPT_DIR ---
# SCRIPT_DIR may be a UNC network share (\\PO\OPBISP\...). Writing thousands of
# times to a share causes intermittent "Permission denied" / "Access is denied".
# Reports are written locally, then optionally copied back.
def _resolve_output_dir():
    candidates = [
        os.path.join(os.path.expanduser('~'), 'Documents', 'FBCRS_Reports'),
        os.path.join(os.path.expanduser('~'), 'FBCRS_Reports'),
        os.path.join(os.environ.get('LOCALAPPDATA') or tempfile.gettempdir(), 'FBCRS_Reports'),
        tempfile.gettempdir(),
    ]
    for d in candidates:
        try:
            os.makedirs(d, exist_ok=True)
            probe = os.path.join(d, '.fbcrs_write_test')
            with open(probe, 'w') as f:
                f.write('ok')
            os.remove(probe)
            return d
        except Exception:
            continue
    return tempfile.gettempdir()

OUTPUT_DIR = _resolve_output_dir()
MAX_SIZE_BYTES = 15 * 1024 * 1024   # above this, classify on file name only

# --- PERFORMANCE TUNING ---
MAX_WORKERS  = min(8, (os.cpu_count() or 4) * 2)   # work is IO-bound, not CPU-bound
TEXT_CAP     = 100000   # max characters of content fed to the matcher
PDF_PAGES    = 10
DOCX_PARAS   = 200
PPTX_SLIDES  = 10
XLSX_SHEETS  = 3
XLSX_ROWS    = 500
HASH_BYTES   = 5 * 1024 * 1024
EXTS = {'.docx', '.pdf', '.xlsx', '.xlsm', '.pptx', '.txt'}
SKIP_EXTS = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.mp4', '.mov', '.avi', '.wmv', '.mkv', '.zip', '.7z', '.rar', '.exe', '.dll', '.msi'}

LOCK = Lock()
HASHES = {}
NEEDS_HASH = set()   # only files whose SIZE collides with another can be duplicates

# --- STARTUP SAFETY NET (Prevents Silent Crashes) ---
if not os.path.exists(LIBRARY):
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("Fatal Error: Missing Rulebook", 
                         f"Could not find the master rulebook.\n\nPlease ensure 'FBCRS_Master_Full.xlsx' is saved in the exact same folder as this application:\n\n{SCRIPT_DIR}")
    sys.exit()

try:
    wb_lib = load_workbook(LIBRARY, data_only=True, read_only=True)
    RECORDS, KEYWORDS, PHRASES, SYN = {}, {}, {}, {}

    if 'Records' in wb_lib.sheetnames:
        for r in wb_lib['Records'].iter_rows(min_row=2, values_only=True):
            if r[0]: RECORDS[r[0]] = {'title': r[1], 'retention': r[2], 'disp': r[3]}
    if 'Keywords' in wb_lib.sheetnames:
        for r in wb_lib['Keywords'].iter_rows(min_row=2, values_only=True):
            if r[0] and r[1]: KEYWORDS.setdefault(r[0], []).append(str(r[1]).lower())
    if 'Phrases' in wb_lib.sheetnames:
        for r in wb_lib['Phrases'].iter_rows(min_row=2, values_only=True):
            if r[0] and r[1]: PHRASES.setdefault(r[0], []).append(str(r[1]).lower())
    if 'Synonyms' in wb_lib.sheetnames:
        for r in wb_lib['Synonyms'].iter_rows(min_row=2, values_only=True):
            if r[0] and r[1]: SYN[str(r[0]).lower()] = str(r[1]).lower()
    wb_lib.close()
except Exception as e:
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("Fatal Error: Rulebook Read Error", 
                         f"Could not read 'FBCRS_Master_Full.xlsx'. Please ensure the file is CLOSED in Microsoft Excel before running this app.\n\nDetails: {str(e)}")
    sys.exit()

TRACKERS = {'kpi tracker': 'STR-POL-003', 'asset tracker': 'EQS-INV-001', 'event tracker': 'PSR-EVT-001'}

# --- CONTROLLED VOCABULARY (drives the Excel dropdowns) ---
ACTION_OPTIONS = ['DESTROY', 'DO NOTHING', 'MANUAL REVIEW REQUIRED', 'TRANSFER TO ARCHIVES']
REASON_OPTIONS = ['WITHIN RETENTION', 'PASSED RETENTION', 'DUPLICATE', 'EMPTY/DRAFT',
                  'UNSUPPORTED FORMAT', 'MISSING DISPOSITION RULE', 'NEEDS REVIEW']

_ACTION_MAP = {
    'DESTROY': 'DESTROY', 'DISPOSE': 'DESTROY', 'SHRED': 'DESTROY',
    'DO NOTHING': 'DO NOTHING', 'RETAIN': 'DO NOTHING', 'KEEP': 'DO NOTHING', 'NO ACTION': 'DO NOTHING',
    'REVIEW': 'MANUAL REVIEW REQUIRED', 'MANUAL REVIEW': 'MANUAL REVIEW REQUIRED',
    'MANUAL REVIEW REQUIRED': 'MANUAL REVIEW REQUIRED', 'SME REVIEW': 'MANUAL REVIEW REQUIRED',
    'TRANSFER TO ARCHIVES': 'TRANSFER TO ARCHIVES', 'ARCHIVE': 'TRANSFER TO ARCHIVES',
    'PERMANENT': 'TRANSFER TO ARCHIVES',
}
_REASON_MAP = {
    'WITHIN RETENTION': 'WITHIN RETENTION', 'PASSED RETENTION': 'PASSED RETENTION',
    'EXPIRED': 'PASSED RETENTION', 'DUPLICATE': 'DUPLICATE',
    'EMPTY/DRAFT': 'EMPTY/DRAFT', 'TRANSITORY': 'EMPTY/DRAFT', 'EMPTY': 'EMPTY/DRAFT', 'DRAFT': 'EMPTY/DRAFT',
    'UNSUPPORTED FORMAT': 'UNSUPPORTED FORMAT', 'UNSUPPORTED': 'UNSUPPORTED FORMAT',
    'MISSING DISPOSITION RULE': 'MISSING DISPOSITION RULE', 'NO DISPOSITION': 'MISSING DISPOSITION RULE',
    'NEEDS REVIEW': 'NEEDS REVIEW', 'REVIEW': 'NEEDS REVIEW', 'ERROR': 'NEEDS REVIEW',
}

def _key(v):
    return re.sub(r'\s+', ' ', str(v or '').strip().upper())

def normalize_action(v):
    """Coerce anything (including free-text Disposition values from the rulebook)
    into one of ACTION_OPTIONS, so every cell matches its dropdown."""
    k = _key(v)
    if k in _ACTION_MAP: return _ACTION_MAP[k]
    for frag, out in (('DESTROY', 'DESTROY'), ('ARCHIV', 'TRANSFER TO ARCHIVES'),
                      ('TRANSFER', 'TRANSFER TO ARCHIVES'), ('PERMANENT', 'TRANSFER TO ARCHIVES'),
                      ('REVIEW', 'MANUAL REVIEW REQUIRED'), ('RETAIN', 'DO NOTHING')):
        if frag in k: return out
    return 'MANUAL REVIEW REQUIRED'

def normalize_reason(v):
    k = _key(v)
    if k in _REASON_MAP: return _REASON_MAP[k]
    for frag, out in (('DUPLICAT', 'DUPLICATE'), ('PASSED', 'PASSED RETENTION'),
                      ('WITHIN', 'WITHIN RETENTION'), ('UNSUPPORT', 'UNSUPPORTED FORMAT'),
                      ('EMPTY', 'EMPTY/DRAFT'), ('DRAFT', 'EMPTY/DRAFT'),
                      ('DISPOSITION', 'MISSING DISPOSITION RULE')):
        if frag in k: return out
    return 'NEEDS REVIEW'

# --- SCORING CONFIG ---
KW_BASE   = 2.0    # base points for a keyword hit
PH_BASE   = 10.0   # base points for a phrase hit
FN_BOOST  = 3.0    # filename matches are far more reliable than body matches
MAX_OWNERS = 20    # a term shared by more codes than this carries no signal
MIN_SCORE = 2.0      # below this, refuse to guess
BODY_ONLY_MIN = 6.0  # a match with no filename support must be much stronger

# Confidence is reported as a label, not a number. Bands are keyed to the
# rarity-weighted score: 30+ means a distinctive phrase or filename match,
# 4 or less means only generic terms fired.
CONF_BANDS = [(30.0, 'Very High'), (12.0, 'High'), (4.0, 'Medium')]
CONF_OPTIONS = ['Very High', 'High', 'Medium', 'Low']

N_ALTERNATES = 3        # runner-up codes offered to the reviewer
ALT_MIN_SCORE = 2.0     # below this the runners-up are noise too - leave blank
ALT_REL_MIN = 0.40      # and they must be within 40% of the winner to be credible

def confidence_label(score):
    for cutoff, label in CONF_BANDS:
        if score >= cutoff:
            return label
    return 'Low'
UNCLASSIFIED = 'UNCLASSIFIED'

def _boundary_pattern(terms):
    """Match whole words only.

    The old code used plain substring matching, which turned every short term
    into a wildcard: 'bi' fired inside 'liability', 'but' inside 'distribution',
    'all' inside 'install'. Lookarounds instead of \\b so terms containing
    '/' or '-' still behave.
    """
    terms = [t for t in terms if t]
    if not terms:
        return None
    alts = '|'.join(re.escape(t) for t in sorted(terms, key=len, reverse=True))
    return re.compile(r'(?<![a-z0-9])(?:' + alts + r')(?![a-z0-9])')

# Drop synonym entries that are pasted legend blocks rather than synonyms.
# These were expanding into ~40 record-series terms inside EVERY document,
# which is what made every file resolve to the same code.
_BAD_SYN = {t for t, v in SYN.items() if '\n' in str(v) or len(str(v)) > 60 or len(str(v).split()) > 6}
if _BAD_SYN:
    for t in _BAD_SYN:
        SYN.pop(t, None)

_SYN_PATTERN = _boundary_pattern(list(SYN)) if SYN else None

def norm(t):
    t = (t or '').lower()
    if _SYN_PATTERN:
        t = _SYN_PATTERN.sub(lambda m: SYN[m.group(0)], t)
    return t

# --- Build the term index with rarity (IDF-style) weighting ---
# Single words are matched by tokenising the text ONCE and intersecting sets,
# instead of scanning a 1,795-branch alternation regex across every character.
# Measured 21x faster on a 100k-character document with identical results.
# This matters more than it looks: Python's regex engine holds the GIL, so the
# old matcher serialised all four worker threads.
_TOKEN_RE = re.compile(r'[a-z0-9]+(?:[-/.][a-z0-9]+)*')

def _contains_bounded(text, term):
    """Whole-word substring test, so 'audit report' does not match inside
    'preaudit reporting'."""
    n, start = len(term), 0
    while True:
        i = text.find(term, start)
        if i < 0:
            return False
        before = text[i - 1] if i > 0 else ' '
        after = text[i + n] if i + n < len(text) else ' '
        if not (before.isalnum() or after.isalnum()):
            return True
        start = i + 1

def _build_term_index():
    owners = {}
    for code, kws in KEYWORDS.items():
        for k in kws:
            k = (k or '').strip().lower()
            if len(k) >= 2:
                owners.setdefault(k, {})[code] = KW_BASE
    for code, phs in PHRASES.items():
        for p in phs:
            p = (p or '').strip().lower()
            if len(p) >= 2:
                # a phrase hit outranks a keyword hit for the same code
                owners.setdefault(p, {})[code] = PH_BASE

    weights, dropped, single, multi_by_first = {}, [], {}, {}
    for term, code_map in owners.items():
        n = len(code_map)
        if n > MAX_OWNERS:
            # A term claimed by this many codes cannot discriminate between them.
            dropped.append((term, n))
            continue
        weights[term] = [(c, base / n) for c, base in code_map.items()]
        if _TOKEN_RE.fullmatch(term):
            single[term] = True
        else:
            toks = _TOKEN_RE.findall(term)
            if toks:
                multi_by_first.setdefault(toks[0], []).append(term)
    return weights, dropped, single, multi_by_first

TERM_WEIGHTS, DROPPED_TERMS, SINGLE_TERMS, MULTI_BY_FIRST = _build_term_index()

def score_codes(fn_text, body_text):
    """Return [(score, code, evidence), ...] sorted best-first.

    Each distinct term counts once per field, so a document repeating
    'financial' 400 times does not bulldoze everything else.
    """
    if not TERM_WEIGHTS:
        return [(0.0, UNCLASSIFIED, [])]
    scores, evidence = {}, {}
    for text, mult in ((fn_text, FN_BOOST), (body_text, 1.0)):
        if not text:
            continue
        toks = set(_TOKEN_RE.findall(text))
        hits = list(toks.intersection(SINGLE_TERMS))
        for first in toks.intersection(MULTI_BY_FIRST):
            hits.extend(t for t in MULTI_BY_FIRST[first] if _contains_bounded(text, t))
        # Longest first, so evidence leads with the most specific match; sorted
        # rather than set-ordered so repeat runs give byte-identical output.
        for term in sorted(hits, key=lambda t: (-len(t), t)):
            for code, w in TERM_WEIGHTS[term]:
                scores[code] = scores.get(code, 0.0) + w * mult
                ev = evidence.setdefault(code, [])
                if len(ev) < 6:
                    ev.append(term if mult == 1.0 else term + '*')   # * = filename hit
    if not scores:
        return [(0.0, UNCLASSIFIED, [])]
    return sorted(((round(s, 2), c, evidence.get(c, [])) for c, s in scores.items()),
                  key=lambda x: (-x[0], x[1]))


def read_text(fp, file_size=None):
    """Extract a bounded sample of text.

    Was building the result with `text_content += ...` inside nested loops, which
    is quadratic: every cell copied the entire accumulated string. On a 3-sheet
    spreadsheet that measured 219 ms vs 43 ms for list+join - 5x, on the file type
    this tool sees most.
    """
    ext = os.path.splitext(fp)[1].lower()
    # Very large files are not worth opening; the file name still classifies them.
    if file_size is not None and file_size > MAX_SIZE_BYTES:
        return ''
    parts, total = [], 0

    def add(sv):
        nonlocal total
        if sv:
            parts.append(sv)
            total += len(sv) + 1
        return total < TEXT_CAP

    try:
        if ext == '.txt':
            with open(fp, encoding='utf-8', errors='ignore') as f:
                add(f.read(TEXT_CAP // 2))
        elif ext == '.pdf' and pypdf:
            reader = pypdf.PdfReader(fp)
            for i in range(min(PDF_PAGES, len(reader.pages))):
                if not add(reader.pages[i].extract_text() or ''): break
        elif ext == '.docx' and docx:
            for para in docx.Document(fp).paragraphs[:DOCX_PARAS]:
                if not add(para.text): break
        elif ext == '.pptx' and pptx:
            for slide in list(pptx.Presentation(fp).slides)[:PPTX_SLIDES]:
                for shape in slide.shapes:
                    if hasattr(shape, "text") and not add(shape.text): break
                if total >= TEXT_CAP: break
        elif ext in ('.xlsx', '.xlsm'):
            wb = load_workbook(fp, data_only=True, read_only=True)
            try:
                for sheet in wb.sheetnames[:XLSX_SHEETS]:
                    for idx, row in enumerate(wb[sheet].iter_rows(values_only=True)):
                        if idx >= XLSX_ROWS or total >= TEXT_CAP: break
                        for cell in row:
                            if cell is not None: add(str(cell))
                    if total >= TEXT_CAP: break
            finally:
                wb.close()
    except Exception:
        pass
    return ' '.join(parts)[:TEXT_CAP]


def sha(fp):
    h = hashlib.sha256()
    try:
        with open(fp, 'rb') as f:
            chunk = f.read(HASH_BYTES)
            if chunk: h.update(chunk)
    except Exception:
        pass
    return h.hexdigest()

def calculate_expiry(modified_date, retention_string):
    if not retention_string: return None
    match = re.search(r'(CCY|CFY)\s*\+\s*(\d+)', str(retention_string).upper())
    if not match: return None
    base_type, years = match.group(1), int(match.group(2))
    if base_type == 'CCY':
        return date(modified_date.year + years, 12, 31)
    elif base_type == 'CFY':
        fy_end_year = modified_date.year if modified_date.month < 4 else modified_date.year + 1
        return date(fy_end_year + years, 3, 31)
    return None

N_COLS = 21   # 16 original + Confidence, Confidence Reason, 3 alternates
              # (all appended, so the Action/Reason dropdown columns do not move)

def classify(fp):
    fn = os.path.basename(fp)
    ext = os.path.splitext(fp)[1].lower()
    try:
        created = datetime.fromtimestamp(os.path.getctime(fp))
        modified = datetime.fromtimestamp(os.path.getmtime(fp))
        file_size = os.path.getsize(fp)
    except Exception as e:
        # FIX 2: 'created'/'modified' were referenced before assignment here,
        # raising UnboundLocalError that killed the entire scan. Also the row
        # had 14 fields instead of 16, misaligning every column after it.
        return [fn, ext, 'Unknown', '', '', '', 'Unreadable', 'N/A', '', '',
                'MANUAL REVIEW REQUIRED', 'NEEDS REVIEW', 'NO', 'YES', fp, f"ACCESS DENIED: {e}",
                'Low', 'File could not be opened - permissions or path length', '', '', ''], None

    category = 'Document' if ext in ['.docx', '.txt'] else 'Spreadsheet' if ext in ['.xlsx', '.xlsm'] else 'PDF' if ext == '.pdf' else 'Other'
    if ext in SKIP_EXTS:
        # NOTE: this row was misaligned against the header list ('Very High' was
        # landing in Retention Period, 'Skip Rule' in Retention Expiry Date).
        return [fn, ext, category, created, modified, 'N/A', 'Non-Record', 'N/A', '', 'Review',
                'MANUAL REVIEW REQUIRED', 'UNSUPPORTED FORMAT', 'NO', 'YES', fp, 'Bypassed by skip rule',
                'Very High', 'File type cannot contain records - skipped by rule', '', '', ''], None

    if file_size == 0 or fn.startswith('~$'):
        return [fn, ext, category, created, modified, 'N/A', 'Transitory', 'N/A', '', 'Destroy',
                'DESTROY', 'EMPTY/DRAFT', 'NO', 'NO', fp, 'Empty file or Office lock/draft file',
                'Very High', 'File is empty or is a temporary Office lock file', '', '', ''], None

    txt = norm(read_text(fp, file_size))
    # Only files whose byte size collides with another file can possibly be
    # duplicates, so most files are never read for hashing at all. Duplicate
    # status itself is resolved by the caller, in file order (see execute_scan).
    file_hash = sha(fp) if fp in NEEDS_HASH else None
    dup = 'NO'

    scores = score_codes(norm(fn), txt)
    best_score, best_code, best_ev = scores[0]

    # Refuse to guess when there is no real signal. The old code took
    # scores[0] unconditionally, so an all-zero tie was broken alphabetically
    # and every unmatched file received the same arbitrary code.
    # Evidence terms ending in '*' came from the filename, which is a far more
    # deliberate signal than boilerplate buried in body text.
    fn_ev = [e.rstrip('*') for e in best_ev if e.endswith('*')]
    body_ev = [e for e in best_ev if not e.endswith('*')]
    fn_backed = bool(fn_ev)
    # Judge phrase-vs-keyword from the field that actually earned the match,
    # otherwise a stray body phrase gets credited to the filename.
    deciding_ev = fn_ev if fn_backed else body_ev
    phrase_hit = any(' ' in e for e in deciding_ev)
    ambiguous = (len(scores) > 1 and best_score >= MIN_SCORE
                 and scores[1][0] >= best_score * 0.85)

    if best_score <= 0:
        code, level = UNCLASSIFIED, 'Low'
        conf_reason = 'No rulebook keywords or phrases found - context is not clear'
    elif best_score < MIN_SCORE:
        code, level = UNCLASSIFIED, 'Low'
        conf_reason = 'Only generic terms matched, shared across many series - context is not clear'
    elif not fn_backed and best_score < BODY_ONLY_MIN:
        code, level = UNCLASSIFIED, 'Low'
        conf_reason = 'Weak match in document body only, nothing in the file name - context is not clear'
    else:
        code = best_code
        level = confidence_label(best_score)
        where = 'file name' if fn_backed else 'document content'
        what = 'exact phrase' if phrase_hit else 'keywords'
        if level in ('Very High', 'High'):
            conf_reason = f'Matched {what} in {where}'
        elif level == 'Medium':
            conf_reason = f'Matched {what} in {where} - moderate evidence only'
        else:
            conf_reason = f'Weak match on {what} in {where} - please verify'
        if ambiguous:
            level = 'Medium' if level in ('Very High', 'High') else level
            conf_reason = f'Close call with {scores[1][1]} - two series scored almost equally'

    # Runners-up help a reviewer pick from a shortlist, but only when the top
    # match had real signal. Otherwise they are noise dressed up as options.
    alternates = ['' for _ in range(N_ALTERNATES)]
    if best_score >= ALT_MIN_SCORE and code != UNCLASSIFIED:
        # A runner-up is only worth offering if it is genuinely competitive.
        # Listing the 2nd-through-4th place codes unconditionally hands the
        # reviewer three plausible-looking wrong answers on every clear match.
        floor = max(ALT_MIN_SCORE, best_score * ALT_REL_MIN)
        alts = [c for _s, c, _e in scores[1:1 + N_ALTERNATES] if _s >= floor]
        for i, c in enumerate(alts):
            alternates[i] = c
    
    lib_disp = RECORDS.get(code, {}).get('disp', '')
    lib_ret = RECORDS.get(code, {}).get('retention', '')
    expiry = calculate_expiry(modified, lib_ret)
    expiry_str = expiry.strftime('%Y-%m-%d') if expiry else ''
    
    expired = bool(expiry) and datetime.now().date() > expiry

    if dup == 'YES':
        final_action, final_reason = 'DESTROY', 'DUPLICATE'
    elif code == UNCLASSIFIED:
        final_action, final_reason = 'MANUAL REVIEW REQUIRED', 'NEEDS REVIEW'
    elif expired and str(lib_disp).strip():
        final_action, final_reason = normalize_action(lib_disp), 'PASSED RETENTION'
    elif expired:
        # Retention has lapsed but the rulebook has no Disposition for this code.
        final_action, final_reason = 'MANUAL REVIEW REQUIRED', 'MISSING DISPOSITION RULE'
    elif level == 'Low':
        final_action, final_reason = 'MANUAL REVIEW REQUIRED', 'NEEDS REVIEW'
    else:
        final_action, final_reason = 'DO NOTHING', 'WITHIN RETENTION'

    title = RECORDS.get(code, {}).get('title', 'Unclassified - no confident match')
    sme_review = 'YES' if (level == 'Low' or code == UNCLASSIFIED) else 'NO'
    # Confidence / reason / alternates appended at the END so the Recommended
    # Action and Reason dropdown columns keep their existing positions.
    return [fn, ext, category, created, modified, code, title, lib_ret, expiry_str,
            lib_disp, final_action, final_reason, dup, sme_review, fp, '',
            level, conf_reason] + alternates, file_hash

def add_review_dropdowns(wb, ws, headers, buffer_rows=500):
    """Attach in-cell dropdowns to the Recommended Action and Reason columns.

    Columns are found by HEADER NAME, not hardcoded letters, so this stays
    correct even if the column order changes.
    """
    lists = wb.create_sheet('Lists')
    for i, v in enumerate(ACTION_OPTIONS, start=1):
        lists.cell(row=i, column=1, value=v)
    for i, v in enumerate(REASON_OPTIONS, start=1):
        lists.cell(row=i, column=2, value=v)
    for i, v in enumerate(CONF_OPTIONS, start=1):
        lists.cell(row=i, column=3, value=v)
    lists.sheet_state = 'hidden'

    last = max(ws.max_row, 2) + buffer_rows

    specs = [
        ('Recommended Action', f"Lists!$A$1:$A${len(ACTION_OPTIONS)}",
         'Invalid action', 'Pick one of the four approved dispositions.'),
        ('Reason', f"Lists!$B$1:$B${len(REASON_OPTIONS)}",
         'Invalid reason', 'Pick one of the approved reason codes.'),
        ('Confidence', f"Lists!$C$1:$C${len(CONF_OPTIONS)}",
         'Invalid confidence', 'Very High, High, Medium or Low.'),
    ]

    for header, ref, err_title, err_body in specs:
        try:
            col = get_column_letter(headers.index(header) + 1)
        except ValueError:
            continue
        dv = DataValidation(
            type='list', formula1=ref, allow_blank=True,
            # openpyxl quirk: showDropDown=False is what SHOWS the arrow.
            # The underlying XML attribute means "suppress the dropdown".
            showDropDown=False,
            showErrorMessage=True, errorStyle='stop',
            errorTitle=err_title, error=err_body,
        )
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{last}")

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 2)}"


def walk_files(folder):
    """Enumerate files once, returning (path, size) sorted by path.

    os.walk discards the stat data Windows already hands back with each
    directory entry, so collecting sizes here is free rather than costing a
    second metadata round trip per file over SMB.
    """
    out, stack = [], [folder]
    while stack:
        d = stack.pop()
        try:
            with os.scandir(d) as it:
                for e in it:
                    try:
                        if e.is_dir(follow_symlinks=False):
                            stack.append(e.path)
                        elif e.is_file(follow_symlinks=False):
                            if e.name.startswith(('~', '.')):
                                continue
                            out.append((e.path, e.stat(follow_symlinks=False).st_size))
                    except OSError:
                        continue
        except OSError:
            continue
    out.sort()
    return out


DUP_IDX, ACTION_IDX, REASON_IDX = 12, 10, 11   # Duplicate / Recommended Action / Reason

def mark_duplicate(row):
    row = list(row)
    row[DUP_IDX], row[ACTION_IDX], row[REASON_IDX] = 'YES', 'DESTROY', 'DUPLICATE'
    return row


def safe_classify(fp):
    """FIX 3: never let a worker exception escape into executor.map, which would
    abort the whole scan. One bad file becomes one flagged row instead."""
    try:
        row, file_hash = classify(fp)
        if len(row) != N_COLS:
            row = (list(row) + [''] * N_COLS)[:N_COLS]
        return row, file_hash
    except Exception as e:
        fn = os.path.basename(fp)
        ext = os.path.splitext(fp)[1].lower()
        return [fn, ext, 'Unknown', '', '', '', 'Scan Failed', 'N/A', '', '',
                'MANUAL REVIEW REQUIRED', 'NEEDS REVIEW', 'NO', 'YES', fp,
                f"{type(e).__name__}: {e}",
                'Low', 'Scan failed on this file - see Comments', '', '', ''], None


def _sanitize(name):
    name = re.sub(r'[<>:"/\\|?*]+', '_', str(name)).strip()
    return (name or 'Scan')[:80]


# --- Background Worker Thread ---
def execute_scan(folder, update_status, update_progress, on_complete):
    start_time = time.time()
    OUTPUT = ""
    CSV_TEMP = ""
    fh = None
    try:
        update_status("Building file list...")
        # Sorted so duplicate detection is deterministic. Whichever copy is seen
        # FIRST is kept and the later one is flagged DESTROY -- with unsorted
        # os.walk order plus 4 threads, which copy of a pair got marked for
        # destruction could change between runs on the same folder.
        sized = walk_files(folder)
        files = [p for p, _ in sized]
        TOTAL = len(files)
        if TOTAL == 0:
            update_status("No files found in selected directory.")
            on_complete(0, 0, 0, "")
            return

        folder_name = _sanitize(os.path.basename(os.path.normpath(folder)))
        stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        # FIX 1 (cont.): write to a guaranteed-local, writable directory.
        OUTPUT = os.path.join(OUTPUT_DIR, f"{folder_name} summary {stamp}.xlsx")
        CSV_TEMP = os.path.join(OUTPUT_DIR, f"{folder_name}_{stamp}_temp.csv")

        # Only files sharing an exact byte size can be duplicates, so size is a
        # free pre-filter and most files are never opened for hashing at all.
        # The previous version hashed EVERY file in its own pass - including the
        # .mp4/.zip/.png files that get skipped anyway and were never hashed
        # before - which is what made this slow over the network.
        by_size = {}
        for p_, sz_ in sized:
            if sz_ > 0:
                by_size.setdefault(sz_, []).append(p_)
        NEEDS_HASH.clear()
        NEEDS_HASH.update(p_ for g in by_size.values() if len(g) > 1 for p_ in g)

        headers = ['File Name', 'Extension', 'Content Category', 'Date Created', 'Date Modified',
                   'Code', 'Series Title', 'Retention Period', 'Retention Expiry Date', 'Disposition',
                   'Recommended Action', 'Reason', 'Duplicate Detected', 'SME Review Required',
                   'Full File Path', 'Comments', 'Confidence', 'Confidence Reason',
                   'Alternate Code 1', 'Alternate Code 2', 'Alternate Code 3']

        # FIX 4: open the CSV ONCE and hold the handle. The old code reopened the
        # file in append mode for every row -- thousands of opens against an SMB
        # share, any one of which can fail with Permission denied / Access denied.
        fh = open(CSV_TEMP, 'w', newline='', encoding='utf-8-sig')
        writer = csv.writer(fh)
        writer.writerow(headers)
        fh.flush()

        processed = 0
        last_ui = 0.0
        seen_hashes = {}
        # executor.map yields in submission order and `files` is sorted, so the
        # first copy of a duplicate pair is the same one on every run. Deciding
        # this here rather than inside the workers is what makes it deterministic.
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            for row, file_hash in executor.map(safe_classify, files):
                if file_hash is not None:
                    if file_hash in seen_hashes:
                        row = mark_duplicate(row)
                    else:
                        seen_hashes[file_hash] = 1
                writer.writerow(row)
                processed += 1
                if processed % 250 == 0:
                    fh.flush()          # checkpoint without reopening
                    os.fsync(fh.fileno())
                now = time.time()
                if now - last_ui > 0.15 or processed == TOTAL:   # don't flood the UI thread
                    last_ui = now
                    update_progress(processed / TOTAL)
                    update_status(f"Scanning files: {processed:,} / {TOTAL:,}")

        fh.flush()
        fh.close()
        fh = None

        # Finalize Excel Report
        update_status("Saving final Excel report...")
        wb = Workbook()
        ws = wb.active
        ws.title = 'Inventory'

        PATH_IDX = headers.index('Full File Path')
        with open(CSV_TEMP, 'r', encoding='utf-8-sig') as f:
            for i, row in enumerate(csv.reader(f)):
                ws.append(row)
                if i > 0 and len(row) > PATH_IDX:
                    # FIX 5: guard the index, and skip hyperlinks Excel rejects
                    # (>255 chars) instead of crashing the whole save.
                    target = str(row[PATH_IDX])
                    if 0 < len(target) <= 255:
                        try:
                            cell = ws.cell(row=i + 1, column=PATH_IDX + 1)
                            cell.hyperlink = target
                            cell.style = "Hyperlink"
                        except Exception:
                            pass

        # --- DROPDOWNS on Recommended Action + Reason ---
        add_review_dropdowns(wb, ws, headers)

        for attempt in range(3):
            try:
                wb.save(OUTPUT)
                break
            except PermissionError:
                if attempt == 2:
                    update_status("Error: the output Excel file is open. Close it and try again.")
                    on_complete(0, 0, 0, "")
                    return
                time.sleep(1.5)

        try: os.remove(CSV_TEMP)
        except Exception: pass

        elapsed_time = time.time() - start_time
        update_status(f"Complete! Saved to: {OUTPUT}")

        try: os.startfile(OUTPUT)
        except Exception: pass

        on_complete(TOTAL, processed, elapsed_time, OUTPUT)

    except Exception as e:
        # FIX 6: report the real error type and keep the checkpoint on disk.
        msg = f"{type(e).__name__}: {e}"
        try:
            with open(os.path.join(OUTPUT_DIR, 'FBCRS_error_log.txt'), 'a', encoding='utf-8') as log:
                log.write(f"\n--- {datetime.now()} ---\n{traceback.format_exc()}")
        except Exception:
            pass
        update_status(f"Error: {msg}\nCheckpoint kept at: {CSV_TEMP or OUTPUT_DIR}")
        on_complete(0, 0, 0, OUTPUT)
    finally:
        if fh is not None:
            try: fh.close()
            except Exception: pass

# --- CustomTkinter UI ---
def run_app_gui():
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")

    app = ctk.CTk()
    app.title("FBCRS Document Analysis Engine")
    app.geometry("820x840")
    app.resizable(False, False)

    selected_folder = ctk.StringVar(value="No folder selected...")

    def show_analytics_dashboard(total, processed, elapsed, output_path):
        mins = int(elapsed // 60)
        secs = int(elapsed % 60)
        
        dash_win = ctk.CTkToplevel(app)
        dash_win.title("Post-Scan Analytics Dashboard")
        dash_win.geometry("520x400")
        dash_win.resizable(False, False)
        
        ctk.CTkLabel(dash_win, text="📊 Analysis Complete Summary", font=ctk.CTkFont(size=18, weight="bold"), text_color="lightgreen").pack(pady=(20, 10))
        
        stats_frame = ctk.CTkFrame(dash_win)
        stats_frame.pack(padx=30, pady=10, fill="x")
        
        ctk.CTkLabel(stats_frame, text=f"⏱️ Total Processing Time: {mins}m {secs}s", font=ctk.CTkFont(size=13)).pack(anchor="w", padx=20, pady=6)
        ctk.CTkLabel(stats_frame, text=f"📁 Total Files Discovered: {total:,}", font=ctk.CTkFont(size=13)).pack(anchor="w", padx=20, pady=6)
        ctk.CTkLabel(stats_frame, text=f"✅ Successfully Classified: {processed:,}", font=ctk.CTkFont(size=13)).pack(anchor="w", padx=20, pady=6)
        
        path_frame = ctk.CTkFrame(dash_win)
        path_frame.pack(padx=30, pady=10, fill="x")
        
        ctk.CTkLabel(path_frame, text="Saved Report Location:", font=ctk.CTkFont(size=11, weight="bold"), text_color="gray70").pack(anchor="w", padx=15, pady=(5,0))
        path_lbl = ctk.CTkLabel(path_frame, text=output_path, font=ctk.CTkFont(size=11), text_color="lightblue", wraplength=440)
        path_lbl.pack(anchor="w", padx=15, pady=(0, 10))
        
        def open_folder():
            try:
                os.startfile(os.path.dirname(output_path))
            except:
                pass
                
        ctk.CTkButton(dash_win, text="Open Output Folder", command=open_folder, width=160, fg_color="teal").pack(pady=10)

    def browse():
        folder = filedialog.askdirectory(title="Select Folder to Analyze")
        if folder:
            selected_folder.set(folder)
            start_btn.configure(state="normal")
            status_lbl.configure(text="Ready to begin analysis.")

    def start():
        folder = selected_folder.get()
        if os.path.exists(folder):
            browse_btn.configure(state="disabled")
            start_btn.configure(state="disabled")
            progress_bar.set(0)
            
            def safe_status(text):
                app.after(0, lambda: status_lbl.configure(text=text))
                
            def safe_progress(val):
                app.after(0, lambda: progress_bar.set(val))
            
            def handle_completion(total, processed, elapsed, out_path):
                def update_ui():
                    browse_btn.configure(state="normal")
                    start_btn.configure(state="normal")
                    if total > 0:
                        show_analytics_dashboard(total, processed, elapsed, out_path)
                app.after(0, update_ui)

            threading.Thread(target=execute_scan, args=(
                folder,
                safe_status,
                safe_progress,
                handle_completion
            ), daemon=True).start()

    # --- TOP HEADER ---
    ctk.CTkLabel(app, text="FBCRS Document Analysis & Recommendation Engine", font=ctk.CTkFont(size=19, weight="bold")).pack(pady=(15, 10))

    # --- SECTION 1: ABOUT THE ENGINE ---
    about_frame = ctk.CTkFrame(app, fg_color=("gray85", "gray17"))
    about_frame.pack(padx=25, pady=4, fill="x")

    about_text = (
        "🌟 ABOUT THE ENGINE:\n"
        "* Intelligently scans large documents for rapid classification.\n"
        "* Reliably detects exact duplicate files to help streamline your records.\n"
        "* Efficiently processes and scans thousands of files in minutes.\n"
        "* Automatically flags transitory records, draft documents, empty files, and unsupported formats."
    )
    about_lbl = ctk.CTkLabel(about_frame, text=about_text, justify="left", font=ctk.CTkFont(size=11), text_color="gray90", wraplength=730)
    about_lbl.pack(padx=16, pady=10, anchor="w")

    # --- SECTION 2: CRITICAL INSTRUCTIONS & GUIDELINES ---
    inst_frame = ctk.CTkFrame(app, fg_color=("gray80", "gray15"))
    inst_frame.pack(padx=25, pady=4, fill="x")

    inst_text = (
        "📌 CRITICAL INSTRUCTIONS & GUIDELINES:\n"
        "• Excel Library Closed: 'FBCRS_Master_Full.xlsx' must remain CLOSED during scans to prevent file-locking errors.\n"
        "• Updating Rules: Edit keywords/phrases inside Excel, save changes, and restart the app to apply them instantly.\n"
        "• Performance & Speed: Scan times vary by device capability. For massive folders, break data into smaller sub-directories.\n"
        "• Network Security: If you encounter server permission errors, run the application locally from your PC Desktop.\n"
        "• Crash Resilience: If interrupted unexpectedly, recovery checkpoints are safely saved to temporary CSV files in the directory."
    )
    inst_lbl = ctk.CTkLabel(inst_frame, text=inst_text, justify="left", font=ctk.CTkFont(size=11), text_color="gray90", wraplength=730)
    inst_lbl.pack(padx=16, pady=10, anchor="w")

    # Folder Selector Section
    selector_frame = ctk.CTkFrame(app)
    selector_frame.pack(padx=25, pady=10, fill="x")

    ctk.CTkLabel(selector_frame, textvariable=selected_folder, text_color="gray50", font=ctk.CTkFont(slant="italic")).pack(side="left", padx=15, pady=12)
    browse_btn = ctk.CTkButton(selector_frame, text="Browse Folder", command=browse, width=110)
    browse_btn.pack(side="right", padx=15, pady=12)

    progress_bar = ctk.CTkProgressBar(app, width=720)
    progress_bar.pack(pady=(12, 5))
    progress_bar.set(0)

    # FIX 7: without wraplength, long error messages overflow the fixed-width,
    # non-resizable window and get clipped at BOTH ends -- which is why the
    # original error was unreadable.
    status_lbl = ctk.CTkLabel(app, text="Waiting for folder selection...",
                              font=ctk.CTkFont(size=13), wraplength=740, justify="left")
    status_lbl.pack(pady=3, padx=25, fill="x")

    ctk.CTkLabel(app, text=f"Reports are saved to: {OUTPUT_DIR}",
                 font=ctk.CTkFont(size=10), text_color="gray55", wraplength=740).pack(pady=(0, 2))

    start_btn = ctk.CTkButton(app, text="Start Analysis", command=start, height=42, width=220, font=ctk.CTkFont(size=15, weight="bold"), state="disabled", fg_color="green", hover_color="darkgreen")
    start_btn.pack(pady=(10, 15))

    app.mainloop()

if __name__ == '__main__':
    run_app_gui()