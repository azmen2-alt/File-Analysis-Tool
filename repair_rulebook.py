r"""
FBCRS Rulebook Repair + Audit
=============================
Repairs the Synonyms sheet (the cause of every file collapsing onto one code)
and prints an audit of remaining data-quality issues.

Usage:  python repair_rulebook.py "path\to\FBCRS_Master_Full.xlsx"

Writes:  <name>_REPAIRED.xlsx   (use this as the new rulebook)
         Records / Keywords / Phrases are copied through UNCHANGED.
"""
import sys, os, re, collections
from openpyxl import load_workbook, Workbook

# Genuine acronym expansions that belong in Synonyms.
GENUINE_SYNONYM_KEYS = {'kpi', 'bi', 'pbi', 'mou', 'ssrs'}


def is_corrupt(term, replacement):
    """A synonym cell holding a pasted multi-line legend, not a synonym."""
    r = str(replacement)
    return ('\n' in r) or len(r) > 60 or len(r.split()) > 6


def salvage_legend(replacement):
    """Corrupt cells begin with the correct expansion, then spill the rest of
    the legend. Keep only the first line."""
    first = str(replacement).replace('\\n', '\n').split('\n')[0]
    return re.sub(r'\s+', ' ', first).strip()


def main(src):
    wb = load_workbook(src, data_only=True)
    out = Workbook()
    out.remove(out.active)

    # --- copy Records / Keywords / Phrases verbatim ---
    for name in ('Records', 'Keywords', 'Phrases'):
        if name not in wb.sheetnames:
            print(f"  WARNING: sheet {name!r} missing from source")
            continue
        ws_in, ws_out = wb[name], out.create_sheet(name)
        for row in ws_in.iter_rows(values_only=True):
            ws_out.append(list(row))

    # --- rebuild Synonyms ---
    syn_out = out.create_sheet('Synonyms')
    syn_out.append(['Term', 'Replacement'])
    legend_out = out.create_sheet('FunctionCodes')      # reference only, engine ignores it
    legend_out.append(['Function Code', 'Meaning', 'Note'])

    kept, moved = [], []
    if 'Synonyms' in wb.sheetnames:
        for r in wb['Synonyms'].iter_rows(min_row=2, values_only=True):
            if not r or not r[0] or r[1] is None:
                continue
            term, repl = str(r[0]).strip().lower(), str(r[1])
            if is_corrupt(term, repl):
                moved.append((term, salvage_legend(repl)))
            elif term in GENUINE_SYNONYM_KEYS or len(term) >= 3:
                kept.append((term, re.sub(r'\s+', ' ', repl).strip().lower()))
            else:
                moved.append((term, salvage_legend(repl)))

    for t, v in kept:
        syn_out.append([t, v])
    for t, v in moved:
        legend_out.append([t.upper(), v, 'Moved out of Synonyms - was corrupting all text matching'])

    dst = os.path.splitext(src)[0] + '_REPAIRED.xlsx'
    out.save(dst)

    print("=" * 68)
    print("SYNONYMS REPAIR")
    print("=" * 68)
    print(f"  kept as real synonyms ({len(kept)}):")
    for t, v in kept:
        print(f"      {t:8} -> {v}")
    print(f"\n  moved to FunctionCodes sheet ({len(moved)}) - these were expanding")
    print("  into whole legend blocks inside every document's text:")
    for t, v in moved:
        print(f"      {t:8} -> {v}")

    # --- advisory audit (no data removed) ---
    kw_owner, ph_owner = collections.defaultdict(set), collections.defaultdict(set)
    if 'Keywords' in wb.sheetnames:
        for r in wb['Keywords'].iter_rows(min_row=2, values_only=True):
            if r[0] and r[1]: kw_owner[str(r[1]).strip().lower()].add(r[0])
    if 'Phrases' in wb.sheetnames:
        for r in wb['Phrases'].iter_rows(min_row=2, values_only=True):
            if r[0] and r[1]: ph_owner[str(r[1]).strip().lower()].add(r[0])

    print("\n" + "=" * 68)
    print("ADVISORY AUDIT (nothing removed - the engine now down-weights these)")
    print("=" * 68)
    worst = sorted(kw_owner.items(), key=lambda x: -len(x[1]))[:10]
    print("  keywords shared by the most codes (near-zero signal value):")
    for k, o in worst:
        print(f"      {len(o):3} codes  {k!r}")
    worst_p = sorted(ph_owner.items(), key=lambda x: -len(x[1]))[:6]
    print("  phrases shared by the most codes:")
    for p, o in worst_p:
        print(f"      {len(o):3} codes  {p!r}")
    singles = sum(1 for o in kw_owner.values() if len(o) == 1)
    print(f"\n  distinctive keywords (owned by exactly 1 code): {singles}/{len(kw_owner)}")

    print(f"\nWROTE: {dst}")
    return dst


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        input("\nPress Enter to exit...")  # Added to prevent instant close when run without file
        sys.exit(1)
    main(sys.argv[1])
    input("\nFinished! Press Enter to exit...")  # Added to prevent instant close after completion